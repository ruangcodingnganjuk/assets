import tkinter as tk
from tkinter import filedialog
import openpyxl
from selenium import webdriver
import pandas as pd
import os

# Buat window utama
window = tk.Tk()
window.title("Program GUI")

# Buat frame untuk kolom "Masukkan file Excel"
frame_file = tk.Frame(window)
frame_file.grid(column=0, row=0)

# Buat label dan entry untuk mengisi file Excel
label_file = tk.Label(frame_file, text="Masukkan file Excel:")
label_file.grid(column=0, row=0, sticky='w')
entry_file = tk.Entry(frame_file)
entry_file.grid(column=1, row=0)

# Buat tombol "Open" untuk membuka file Excel
def open_file():
  file = tk.filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
  entry_file.delete(0, tk.END)
  entry_file.insert(0, file)
button_open = tk.Button(frame_file, text="Open", command=open_file)
button_open.grid(column=2, row=0, sticky='e')

# Buat frame untuk kolom "Pilih sheet"
frame_sheet = tk.Frame(window)
frame_sheet.grid(column=0, row=1)

# Buat label dan menu dropdown untuk memilih sheet
label_sheet = tk.Label(frame_sheet, text="Pilih sheet:")
label_sheet.grid(column=0, row=0, sticky='w')
file = entry_file.get()

# cek jenis file
_, file_extension = os.path.splitext(file)

if file_extension in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
    wb = openpyxl.load_workbook(file)
    # ... code untuk mengambil sheet range dan mengisi form
else:
    df = pd.read_excel(file)
    # ... code untuk mengambil sheet range dan mengisi form
    
sheet_var = tk.StringVar(window)
sheet_dropdown = tk.OptionMenu(frame_sheet, sheet_var, *wb.sheetnames)
sheet_dropdown.grid(column=1, row=0)

# Buat frame untuk kolom "Pilih sheet range"
frame_range = tk.Frame(window)
frame_range.grid(column=0, row=2)

# Buat label dan spinbox untuk memilih sheet range
label_range = tk.Label(frame_range, text="Pilih sheet range:")
label_range.grid(column=0, row=0, sticky='w')
range_var = tk.StringVar(window)
ws = wb[sheet_var.get()]
range_spinbox = tk.Spinbox(frame_range, from_=1, to=ws.max_row, textvariable=range_var)
range_spinbox.grid(column=1, row=0)

# Buat frame untuk kolom "Masukkan link Google Form"
frame_link = tk.Frame(window)
frame_link.grid(column=0, row=3)

# Buat label dan entry untuk mengisi link Google Form
label_link = tk.Label(frame_link, text="Masukkan link Google Form:")
label_link.grid(column=0, row=0, sticky='w')
entry_link = tk.Entry(frame_link)
entry_link.grid(column=1, row=0)

# Buat frame untuk elemen-elemen form
frame_elements = tk.Frame(window)
frame_elements.grid(column=0, row=4)

# Buat list untuk menyimpan elemen-elemen form
elements = []

# Buat tombol untuk menambah elemen
def add_element():
    # Buat frame untuk elemen baru
    element_frame = tk.Frame(frame_elements)
    element_frame.grid(column=0, row=len(elements)+1)
    # Buat label dan entry untuk elemen baru
    element_label = tk.Label(element_frame, text="Elemen {}:".format(len(elements)+1))
    element_label.grid(column=0, row=0, sticky='w')
    element_entry = tk.Entry(element_frame)
    element_entry.grid(column=1, row=0)
    # Tambahkan elemen ke list
    elements.append((element_label, element_entry))
button_add = tk.Button(text="+", command=add_element)
button_add.grid(column=2, row=4, sticky='w')

# Buat tombol untuk menjalankan program
def generate():
    # Baca file Excel dan sheet yang dipilih
    file = entry_file.get()
    sheet = sheet_var.get()
    wb = openpyxl.load_workbook(file)
    ws = wb[sheet]
    # Baca sheet range yang dipilih
    sheet_range = range_var.get()
    # Baca link Google Form
    link = entry_link.get()
    # Inisialisasi browser menggunakan Selenium
    driver = webdriver.Chrome()
    # Buka link Google Form
    driver.get(link)
    # Isi elemen-elemen form sesuai dengan data di sheet Excel
    for row in ws.iter_rows(min_row=2, max_row=int(sheet_range)):
          for element, cell in zip(elements, row):
              label, entry = element
              # Ambil nama elemen dari label
              name = label['text'].lower()
              # Cari elemen di form dengan nama yang sesuai
              elem = driver.find_element_by_name(name)
              # Isi elemen dengan data dari sheet Excel
              elem.send_keys(cell.value)
              # Kirim form
    driver.find_element_by_xpath('//*[@id="mG61Hd"]/div/div[2]/div[3]/div[1]/div/div/content/span').click()
    # Tutup browser
    driver.close()

# Buat tombol untuk menjalankan program
button_generate = tk.Button(text="Generate", command=generate)
button_generate.grid(column=1, row=4, sticky='w')

# Jalankan window utama
window.mainloop()

