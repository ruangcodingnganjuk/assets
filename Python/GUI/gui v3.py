import tkinter as tk
from tkinter import filedialog
import openpyxl
from selenium import webdriver

# Buat window utama
window = tk.Tk()
window.title("ISI PINTAR. GAK PINTAR YOWES")

# Buat frame untuk kolom "Masukkan file Excel"
frame_file = tk.Frame(window)
frame_file.grid(column=0, row=0)

# Buat label dan entry untuk mengisi file
label_file = tk.Label(frame_file, text="Masukkan file Excel:")
label_file.grid(column=0, row=0, sticky='w')
entry_file = tk.Entry(frame_file)
entry_file.grid(column=1, row=0)

# Buat tombol "Open" untuk membuka file dari komputer
def open_file():
  file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
  entry_file.delete(0, tk.END)
  entry_file.insert(0, file)
  # Baca daftar sheet dari file Excel
  wb = openpyxl.load_workbook(file)
  sheets = wb.sheetnames
  # Kosongkan daftar sheet jika sebelumnya sudah ada isian
  sheet_var.set('')
  sheet_menu['menu'].delete(0, 'end')
  # Tambahkan daftar sheet ke dalam menu
  for sheet in sheets:
    sheet_menu['menu'].add_command(label=sheet, command=tk._setit(sheet_var, sheet))

button_open = tk.Button(frame_file, text="Open", command=open_file)
button_open.grid(column=2, row=0, sticky='e')

# Buat frame untuk kolom "Pilih sheet"
frame_sheet = tk.Frame(window)
frame_sheet.grid(column=0, row=1)

# Buat label dan menu untuk memilih sheet
label_sheet = tk.Label(frame_sheet, text="Pilih sheet:")
label_sheet.grid(column=0, row=0, sticky='w')
sheet_var = tk.StringVar(window)
sheet_menu = tk.OptionMenu(frame_sheet, sheet_var, '')
sheet_menu.grid(column=1, row=0)

# Buat frame untuk kolom "Masukkan link Google Form"
frame_link = tk.Frame(window)
frame_link.grid(column=0, row=2)

# Buat label dan entry untuk mengisi link Google Form
label_link = tk.Label(frame_link, text="Masukkan link Google Form:")
label_link.grid(column=0, row=0, sticky='w')
entry_link = tk.Entry(frame_link)
entry_link.grid(column=1, row=0)

# Buat list untuk menyimpan elemen yang akan diisi
elements = []

# Buat frame untuk elemen
frame_elements = tk.Frame(window)
frame_elements.grid(column=0, row=3)

# Buat fungsi untuk menambah elemen baru ke dalam list


def add_element():

    # Buat label dan entry baru
    label = tk.Label(frame_elements, text="Elemen:")
    label.grid(column=0, row=len(elements), sticky='w')
    entry = tk.Entry(frame_elements)
    entry.grid(column=1, row=len(elements))
    # Tambahkan ke dalam list
    elements.append((label, entry))

# Buat tombol untuk menambah elemen
label = tk.Label(frame_elements, text='element')
label.grid(column=0, row=len(elements), sticky='w')
button_add = tk.Button(frame_elements, text="+", command=add_element)
button_add.grid(column=2, row=0, sticky='n')
  

# Buat tombol "Generate" untuk memproses data dan mengisi form
def generate():
    # Baca file Excel dan sheet yang dipilih
    file = entry_file.get()
    sheet = sheet_var.get()
    wb = openpyxl.load_workbook(file)
    ws = wb[sheet]

    # Baca link Google Form
    link = entry_link.get()
    # Inisialisasi browser menggunakan Selenium
    driver = webdriver.Chrome()
    driver.get(link)

    # Proses setiap baris di sheet Excel
    for row in ws.iter_rows(min_row=2):
    # Isi elemen-elemen form sesuai dengan data di sheet Excel
      for element, cell in zip(elements, row):
        label, entry = element
        # Ambil nama elemen dari label
        name = label['text'].lower()
        # Cari elemen di form dengan nama yang sesuai
        elem = driver.find_element_by_xpath(name)
        # Isi elemen dengan data dari sheet Excel
        elem.send_keys(cell.value)
#   # Kirim form
#     driver.find_element_by_xpath('//*[@id="mG61Hd"]/div/div[2]/div[3]/div[1]/div/div/content/span').click()
# Tutup browser
    driver.close()

button_generate = tk.Button(text="Generate", command=generate)
button_generate.grid(column=1, row=4, sticky='w')

# Jalankan window utama
window.mainloop()


